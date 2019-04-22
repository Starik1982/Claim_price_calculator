import tkinter as tk
from tkinter import *
from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup


from tkinter import messagebox
from datetime import datetime


class MyApp:
    list_ttn_for_screen = ''    # Used for showing ttn list on screen
    list_ttn = []               # Used for saving ttn list(name, date, ttn sum, percents, inflation index, inflation sum)
    iis = 0

    def __init__(self,  master):
        self.master = master


        '''''''''''''''''screen section'''''''''''''''''
        self.screen = Label(bg='#ccc')
        self.screen.place(relx=.5, rely=.1, anchor="c", height=240, width=680)
        self.screen['text'] = 'Список накладних.'


        '''''''''''''''''ttn section'''''''''''''''''
        self.ttn_name = StringVar(value='документ')
        self.ttn_date = StringVar(value='дд.мм.рррр')
        self.ttn_sum = StringVar(value='сума')
        self.percents = StringVar(value='штраф у %')

        self.ttn_name_entry = Entry(textvariable=self.ttn_name)
        self.ttn_date_entry = Entry(textvariable=self.ttn_date)
        self.ttn_sum_entry = Entry(textvariable=self.ttn_sum)
        self.percents_entry = Entry(textvariable=self.percents)

        self.ttn_name_entry.place(relx=.1, rely=.5, anchor="c")
        self.ttn_date_entry.place(relx=.3, rely=.5, anchor="c")
        self.ttn_sum_entry.place(relx=.5, rely=.5, anchor="c")
        self.percents_entry.place(relx=.7, rely=.5, anchor="c")

        self.message_button = Button(text="додати документ", command=self.ttn_list)
        self.message_button.place(relx=.9, rely=.5, anchor="c")


        '''''''''''''''''ttn section'''''''''''''''''

        # self.file_name = StringVar()
        # self.file_name_entry = Entry(textvariable=self.file_name)
        # self.file_name_entry.place(relx=.5, rely=.6, anchor="c")

        self.message_button = Button(text="розрахунок", command=self.start)
        self.message_button.place(relx=.5, rely=.75, height = 100, width = 550, anchor="c")

        # self.message_button = Button(text="TEST", command=self.get_nbu)
        # self.message_button.place(relx=.7, rely=.8, anchor="c")

    def ttn_list(self):
        a = self.ttn_name.get()
        b = self.ttn_date.get()
        c = self.ttn_sum.get()
        d = self.percents.get()
        if ',' in c:
            c = c.replace(',', '.')
        try:
            test = float(c)
        except:
            self.clickAbout('\n\n\n\nУ поле "сума" можна ввести\n тільки ціле або дробове число.')
            raise

        if ',' in d:
            d = d.replace(',', '.')

        try:
            test1 = float(d)
        except:
            self.clickAbout('\n\n\n\nУ поле "штраф у %" можна ввести\n тільки ціле або дробове число.')
            raise

        try:
            split = b.split('.')
            day = int(split[0])
            month = int(split[1])
            yar = int(split[2])
            date = datetime(yar, month, day)
            self.list_ttn.append([a,date,c,d])
            self.list_ttn_for_screen = self.list_ttn_for_screen +('ТТН № ' + a + ' від ' +  b + ' на загальну суму ' + c + ' грн. \n')
            self.screen_out(self.list_ttn_for_screen)
        except:
            self.clickAbout('\n\n\n\nНевірно вказана дата документа.\n Введіть дату у форматі "ДД.ММ.РРРР".')

    def clickAbout(self, ABOUT_TEXT):
        toplevel = Toplevel()
        toplevel.geometry('250x200+500+150')
        toplevel.title('Помилка!')
        label1 = Label(toplevel, text=ABOUT_TEXT, anchor="c")
        label1.place(x=0.1, y=0.1)
        label1.pack()



    def screen_out(self, x):
        self.screen['text'] = x

    def start(self):
        self.inflation_calculation()
        self.double_discount_rate_nbu()
        self.create_document_exel()
        self.percents_ttn()


    def create_document_exel(self):
        wb = Workbook()
        ws = wb.active
        ws['B2'] = 'РОЗРАХУНОК ЗАБОРГОВАНОСТІ ТА ШТРАФНИХ САНКЦІЙ'
        ws['A3'] = 'Дебіторська заборгованість Позивача розраховується, як сума вартості тоовару поставленого за всіми неоплаченими ТТН'
        ws.append([str(self.str_sum_of_debt()) + str(self.sum_of_debt()) + ' грн.'])
        ws.append([])
        ws.append([])
        ws.append(['ІНФЛЯЦІЙНІ ВТРАТИ'])
        ws.append(['Загальний індекс інфліції за період заборгованості по одній товарно-транспортній накладінй розраховується за формулою:'])
        ws.append(['ІІС = ( ІІ1 : 100 ) * ( ІІ2 : 100 ) * ( ІІ3 : 100 ) * ... ( ІІZ : 100 );'])
        ws.append(['ІІ1 - індекс інфляції за перший місяць прострочення;'])
        ws.append(['ІІZ - індекс інфляції за останній місяць прострочення;'])
        ws.append(['ІІC - індекс інфляції за період прострочення.'])
        ws.append([])

        for i in self.list_ttn:
            ws.append(['ІІC для ТТН № ' + i[0] + ' від ' + str(i[1].date()) + ' року  за період з ' + str(i[1].date())+
                       ' по ' + str(datetime.now().date()) +' дорівнює:' ])
            formula = ''
            iteration = 0
            for y in self.get_html():
                if y[0] >= i[1]:
                    formula = formula + str(y[1]) + '/100 * '
                iteration += 1
                if iteration == 8:
                    if formula != '':
                        ws.append([formula])
                    formula = ''
                    iteration = 0
            ws.append([formula[:-2] + ' = ' + str(i[4])])
            ws.append(['Інфляційне збільшення по ТТН № '+ i[0] +' від ' + str(i[1].date()) + ' складає:'])
            ws.append([i[2]+'грн. *'+str(i[4])+'-'+i[2]+'грн. ='+str(i[5])+ 'грн.'])
            ws.append([])
        ws.append(['Пеня'])
        ws.append([])
        ws.append(['Розрахунок суми пені здійснюється за формулою:'])
        ws.append(['Пеня = С x 2УСД x Д : 100, де '])
        ws.append(['С - сума заборгованості за період,'])
        ws.append(['2 УСД - подвійна облікова ставка НБУ в день прострочення'])
        ws.append(['Д - кількість днів прострочення'])
        ws.append([])
        for i in self.list_ttn:
            ws.append(['Пеня для ТТН № ' + i[0] + ' від ' + str(i[1].date()) + ' сумою боргу ' + str(i[2]) + ' року  за період з ' + str(i[1].date())+
                       ' по ' + str(datetime.now().date()) +' дорівнює:' ])
            ws.append([])
            first_day_period = i[1]
            index = self.first_index(i[1])
            for y in self.get_nbu():
                if y[0] > first_day_period:
                    days = y[0] - first_day_period
                    days = str(days)
                    days = int(days.partition(' ')[0])
                    rethult = float(i[2]) * (index * 2 / 365) * days / 100
                    index = y[1]
                    ws.append(['За період з '+ str(first_day_period.date()) +' по ' + str(y[0].date()) +
                               ' при обліковій ставці НБУ ' + str(y[1]) + ' розмір пені складає ' + str(round(rethult, 2)) + ' грн.'])
                    first_day_period = y[0]
            ws.append(['Загальна пеня по накладній № ' + i[0] + ' від ' + str(i[1].date()) + ' складає ' + str(round(i[6], 2)) + ' грн.' ])
            ws.append([])
        ws.append([])
        ws.append(['РОЗРАХУНОК ШТРАФУ'])
        ws.append(['Розрахунок відбувається за формулою: штраф = сума боргу × розмір штрафних відсотків / 100'])
        ws.append([])
        for i in self.list_ttn:
            days = datetime.now() - i[1]
            days = str(days).partition(' ')[0]
            ws.append(['Для накладної № ' + i[0] + ' від ' + str(i[1].date()) + ' розмір штрафу складає:'])
            ws.append(['Сума боргу ' + i[2] + ' * розмір штрафних відсотків ' + i[3]+ '% / 100  = ' + str(round(i[6], 2)) + ' грн.'])
            ws.append([])
        wb.save('Розрахунок заборгованості ' + str(datetime.now().date()) + ".xlsx")

    def sum_of_debt(self):
        x = 0
        for i in self.list_ttn:
            x = x + round(float(i[2]), 2)
        return x

    def str_sum_of_debt(self):
        x = ''
        for i in self.list_ttn:
            x = x + ' + ' + i[2]
        x = x + ' = '
        x = x[2::]
        return x

    def get_html(self):
        yea = 2000
        month = 1
        day = 1
        inflation_index = []
        url = 'https://index.minfin.com.ua/economy/index/inflation/'
        r = requests.get(url)
        soup = BeautifulSoup(r.text, 'lxml')
        ads = soup.find('div', class_='compact-table' ).find_all('tr')[1::]
        for i in ads:
            x = i.find_all('td')
            for y in x:
                if month > 12:
                    yea += 1
                    month = 1
                date = datetime(yea, month, day)
                month += 1
                yy = y.text.strip()
                if yy !='':
                    inflation_index.append([date, yy.replace(',', '.')])

        return inflation_index


    def inflation_calculation(self):
        itr = 0
        now = datetime.now()
        index_inflation = self.get_html()
        for i in self.list_ttn:
            ttn_date = i[1]
            index = 1
            for y in index_inflation:
                if ttn_date <= y[0]:
                    ii_sum = float(y[1]) / 100 * index
                    index = ii_sum
            self.list_ttn[itr].append(index)
            itr+=1
        for i in self.list_ttn:
            sum = float(i[2])
            result = sum * i[4] - sum
            i.append(round(result, 2))


    def get_nbu(self):
        all_index_nbu = []
        date_index_nbu = []
        index_nbu = []
        url = 'https://index.minfin.com.ua/banks/nbu/refinance/'
        r = requests.get(url)
        soup = BeautifulSoup(r.text, 'lxml')
        ads = soup.find('div', id='idx-wrapper').find('table').find_all('tr')
        for i in ads:
            item_1 = i.find_all('td')[:-2]
            item_2 = i.find_all('td')[1:-1]
            for y in item_1:
                date = y.text.strip()[2:12]
                day = int(date[0:2])
                month = int(date[3:5])
                year = int(date[6:])
                date = datetime(year, month, day)
                date_index_nbu.append(date)
            for y in item_2:
                x = y.text.strip()
                index = float(x.replace(',', '.'))
                index_nbu.append(index)
        for i in range(0, len(index_nbu)):
            in_list = []
            in_list.append(date_index_nbu[i])
            in_list.append(index_nbu[i])
            all_index_nbu.append(in_list)
            in_list = []
        a = len(all_index_nbu)
        in_list.append(datetime.now())
        in_list.append(all_index_nbu[a-1][1])
        all_index_nbu.append(in_list)
        return all_index_nbu

    def first_index(self, ttn_date):
        iterethion = 0
        get_nbu = self.get_nbu()
        for y in get_nbu:
            if ttn_date>y[0]:
                iterethion +=1
        return get_nbu[iterethion-1][1]

    def double_discount_rate_nbu(self):
        sum = 0
        iteretions = 0
        ttn = self.list_ttn
        get_nbu = self.get_nbu()
        for i in ttn:
            index = self.first_index(i[1])
            period = i[1]
            for y in get_nbu:
                if period < y[0]:
                    days = y[0] - period
                    days = str(days)
                    days = int(days.partition(' ')[0])
                    rethult = float(i[2])*(index*2/365)*days/100
                    period = y[0]
                    index = y[1]
                    sum = sum + rethult
            self.list_ttn[iteretions].append(sum)
            iteretions += 1


    def percents_ttn(self):
        iteretions = 0
        ttn_list = self.list_ttn
        for i in ttn_list:
            percents = float(i[2]) * int(i[3]) / 100
            self.list_ttn[iteretions].append(percents)
            iteretions += 1


def main():
    root = tk.Tk()
    root.title("Розрахунок ціни позову")
    root.geometry('700x350+250+100')
    app = MyApp(root)
    root.mainloop()

if __name__ == '__main__':
    main()

input()
